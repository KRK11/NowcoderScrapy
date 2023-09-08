import re
import sys
import queue
import requests
import user_agent
from os import path
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from openpyxl import Workbook, load_workbook


class spider_main:
    # 初始化工作,页数，文件名
    def __init__(self, num):
        self.team = queue.Queue()
        self.member = []
        self.num = int(num)
        self.count = 2
        self.exists = False
        self.Date_Rank = 'Data_Rank\Rank' + str(self.num) + '.xlsx'
        self.Url_Member = 'Url_Member\Member' + str(self.num) + '.txt'
        self.Date_Member = 'Data_Member\Member' + str(self.num) + '.xlsx'
        self.url = 'https://ac.nowcoder.com/acm/contest/rating-index?pageSize=50&searchUserName=&onlyMyFollow=false&page={}'
        self.pre = 'https://ac.nowcoder.com'
        self.suf = '/practice-coding'
        self.head = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en',
            "User-Agent": user_agent.generate_user_agent()
        }
        self.rank1 = {}
        self.rank2 = {}
        self.rank3 = {}
        self.member1 = {}

        # 判重
        try:
            if path.exists(self.Date_Rank) and self.num != 1:
                self.exists = True
                return
        except Exception:
            print('Error on line ', sys._getframe().f_lineno, '!', sep='')
            pass

        # 创建空Rank.xlsx
        try:
            wb = Workbook()
            wb.create_sheet('Rank')
            rank = wb['Rank']
            rank = wb['Rank']
            rank.merge_cells('A1:Q1')
            rank['A1'] = 'Rank'
            rank['A1'].alignment = Alignment(horizontal='center', vertical='center')
            rank['A2'] = 'rank'
            rank['B2'] = 'name'
            rank['C2'] = 'school'
            rank['D2'] = 'description'
            rank['E2'] = 'rating'
            rank['F2'] = 'id'
            rank['G2'] = 'rating contest'
            rank['H2'] = 'contest'
            rank['I2'] = 'fans'
            rank['J2'] = 'member num'
            rank['K2'] = 'challenge'
            rank['L2'] = 'accept'
            rank['M2'] = 'summit'
            rank['N2'] = 'ac_rate'
            rank.merge_cells('O2:Q2')
            rank['O2'] = 'member'
            rank['O2'].alignment = Alignment(horizontal='center', vertical='center')
            wb.save(self.Date_Rank)
        except Exception:
            print('Error on line ', sys._getframe().f_lineno, '!', sep='')
            pass

        # 创建空Member_link
        try:
            with open(self.Url_Member, 'w', encoding='utf-8'):
                pass
        except Exception:
            print('Error on line ', sys._getframe().f_lineno, '!', sep='')
            pass

        # 创建空Member.xlsx
        try:
            wb = Workbook()
            wb.create_sheet('Information')
            info = wb['Information']
            info['A1'] = 'Information'
            info.merge_cells('A1:K1')
            info['A1'].alignment = Alignment(horizontal='center', vertical='center')
            info['A2'] = 'name'
            info['B2'] = 'rating'
            info['C2'] = 'rank'
            info['D2'] = 'rating_contest'
            info['E2'] = 'contest'
            info['F2'] = 'attention'
            info['G2'] = 'fans'
            info['H2'] = 'challenge'
            info['I2'] = 'accept'
            info['J2'] = 'summit'
            info['K2'] = 'ac_rate'
            wb.save(self.Date_Member)
        except Exception:
            print('Error on line ', sys._getframe().f_lineno, '!', sep='')
            pass

    # 根据排名将团队排名信息导入excel
    def save_data(self, bs_tag,id):
        if not bs_tag.find_all('i',{'class':'icon-nc-member'}):
            return
        self.rank1[id] = []
        span = bs_tag.find_all('span')
        self.rank1[id].append(' ' + span[0].text)
        self.rank1[id].append(' ' + span[1].text)
        state = span[2].find('a')
        if state: self.rank1[id].append(' ' + state.text)
        else: self.rank1[id].append(' 暂无')
        self.rank1[id].append(' ' + span[3].text)
        self.rank1[id].append(' ' + span[4].text)
        return

    # 网页解析保存排名与团队首页链接
    def get_url(self, html):
        bs = BeautifulSoup(html.text, 'lxml')
        rank = bs.find_all('tr')
        for num in range(1, len(rank)):
            user = rank[num].find('a').get('href')
            self.team.put(self.pre + user)
            self.save_data(rank[num], user[21:].strip())
        return

    # 获取队伍成员名字与链接
    def get_member(self, id):
        try:
            url = 'https://ac.nowcoder.com/acm/team/member-list?token=&teamId=' + id
            html = requests.get(url, headers=self.head)
            self.rank3[id] = re.findall(r'"name":"(.{1,30})",', html.text)  # member_name
            self.rank3[id] = [' ' + item for item in self.rank3[id]]
            member_uid = re.findall(r'"uid":(\d+),', html.text)  # member_uid
            for member in member_uid:
                link = 'https://ac.nowcoder.com/acm/contest/profile/' + member
                if link not in self.member:
                    self.member.append(link)
        except Exception:
            print('Error on line ', sys._getframe().f_lineno, '!', sep='')
            pass
        return

    # 导入队伍信息页
    def get_data1(self, url,id):
        try:
            html = requests.get(url, headers=self.head)
            bs = BeautifulSoup(html.text, 'lxml')

            # 如果为个人则列入队列中
            belong = bs.find('span', {'class': 'crumbs-end font-green'}).text

            if belong == '个人主页':
                if url not in self.member:
                    self.member.append(url)
            else:
                state = bs.find_all('div', {'class': 'state-num'})
                self.rank2[id] = [id,' ' + state[2].text,' ' + state[3].text]
                state = bs.find_all('div', {'class': 'status-item'})
                self.rank2[id].extend([' ' + state[2].find('a').text,' ' + state[3].find('a').text])

                # 导入队伍训练信息
                url = url + self.suf
                html = requests.get(url, headers=self.head)
                bs = BeautifulSoup(html.text, 'lxml')
                state = bs.find_all('div', {'class': 'state-num'})
                self.rank2[id].extend([' ' + state[i].text for i in range(4)])

                self.get_member(id)

        except Exception:
            print('Error on line ', sys._getframe().f_lineno, '!', sep='')
            pass
        return

    # 导入队员信息
    def get_data2(self, url,id):
        try:
            html = requests.get(url, headers=self.head)
            bs = BeautifulSoup(html.text, 'lxml')
            self.member1[id] = [' ' + bs.find('a', {'class': 'coder-name'}).text] #name
            state = bs.find_all('div', {'class': 'state-num'})
            self.member1[id].extend([' ' + state[i].text for i in range(4)]) #rating,rank,rcon,con
            state = bs.find_all('div', {'class': 'status-item'})[2].find('div').find_all('a')
            self.member1[id].extend([' ' + state[i].text for i in range(2)]) #attention,fans
            if self.member1[id][2] ==' 暂无' or self.member1[id][2] ==' 9999+':
                self.member1[id][2] = ' 10000000'
            # 导入队员训练信息
            html = requests.get(url + self.suf, headers=self.head)
            bs = BeautifulSoup(html.text, 'lxml')
            state = bs.find_all('div', {'class': 'state-num'})
            self.member1[id].extend([' ' + state[i].text for i in range(4)]) #challenge,accept,summit,ac_rate

        except Exception:
            print('Error on line ', sys._getframe().f_lineno, '!', sep='')
            pass
        return

    # 爬虫主体
    def spider(self):
        if self.exists:
            print('Successfully import!')
            return

        # 排名前self.num*50的队伍前部信息
        for num in range(1, self.num + 1):
            try:
                html = requests.get(self.url.format(num), headers=self.head)
            except Exception:
                print('Error on line ', sys._getframe().f_lineno, '!', sep='')
                continue
            self.get_url(html)

        # 队伍后半信息，队员链接
        self.count = 2
        while not self.team.empty():
            try:
                url = self.team.get()
                self.count = self.count + 1
                self.get_data1(url, url[44:].strip())
                # print('Successfully import the ' + str(self.count - 2) + 'th team!')
                print('Team_now:', format((self.count - 2) / (self.num * 50) * 100, '.3f'), '%')
            except Exception:
                print('Error on line ', sys._getframe().f_lineno, '!', sep='')
                continue

        # 保存队伍数据
        try:
            wb = load_workbook(self.Date_Rank)
            rank = wb['Rank']
            self.rank1 = sorted(self.rank1.items(),key = lambda x:eval(x[1][0]))
            self.rank1 = {i[0]:i[1] for i in self.rank1}
            for id in self.rank1:
                rk = self.rank1[id]
                if id in self.rank2:
                    rk.extend(self.rank2[id])
                if id in self.rank3:
                    rk.extend(self.rank3[id])
                if len(rk)>=17:
                    rank.append(rk)
            wb.save(self.Date_Rank)
        except Exception:
            print('Error on line ', sys._getframe().f_lineno, '!', sep='')
            return

        # 目前队伍读取完毕！

        # 存储队员链接
        try:
            with open(self.Url_Member, 'a', encoding='utf-8') as fp:
                for link in self.member:
                    fp.write(link + '\n')
        except Exception:
            print('Error on line ', sys._getframe().f_lineno, '!', sep='')
            pass

        # 处理队员链接
        self.count = 2
        for url in self.member:
            self.count = self.count + 1
            self.get_data2(url,url[44:].strip())
            # print('Successfully import the ' + str(self.count - 2) + 'th member!')
            print('Member_now:', format((self.count - 2) / len(self.member) * 100, '.3f'), '%')

        # 保存队员数据
        try:
            wb = load_workbook(self.Date_Member)
            info = wb['Information']
            self.member1 = sorted(self.member1.items(), key=lambda x: eval(x[1][2]))
            self.member1 = {i[0]: i[1] for i in self.member1}
            for value in self.member1.values():
                if len(value)>=11:
                    info.append(value)
            wb.save(self.Date_Member)
        except Exception:
            print('Error on line ', sys._getframe().f_lineno, '!', sep='')
            return

        print('Successfully import!')
        return
