import os
from os.path import join
import numpy as np
from scipy.optimize import leastsq
from openpyxl import load_workbook
from plotly.offline import plot
import plotly.graph_objs as go
from plotly.subplots import make_subplots


# 最终运行类
class FINAL:
    try:
        team_rank = []
        team_rating = []
        team_school = []
        team_rating_contest = []
        team_contest = []
        team_member = []
        team_ac_summit = []
        team_ac_rate = []
        member_rank = []
        member_rating = []
        member_rating_contest = []
        member_contest = []
        member_ac_summit = []
        member_ac_rate = []
    except Exception:
        print('ERROR!')

    #初始化询问
    def __init__(self):
        print('*' * 100)
        print('输入爬虫方式：')
        print('采用非框架式的request的bs4匹配模式：1')
        print('采用框架scrapy的xpath匹配模式：2')
        print('*' * 100)
        self.option = int(input('输入选择方式：'))
        while self.option < 1 or self.option > 2:
            self.option = int(input('输入错误！重新输入：'))
        print('*' * 100)
        print('查找范围（1~200）页，每页50队/人')
        print('注意范围为10时，request模式下开发者测评时间为800s，scrapy模式下测试时间385s')
        print('由于爬取页面过多,输入范围过大时请耐心等待！')
        print('数据集加载完后将不会重新爬取，若需要重新爬取，请将对应数据删掉')
        print('为了方便测试，数据集已经加载了2，10，20，50，100的数据')
        print('*' * 100)
        self.num = input('输入查询页数（当前页数及前面的所有页数）：').strip()
        while int(self.num) < 1 or int(self.num) > 200:
            self.num = input('输入错误！重新输入：').strip()
        self.adr = os.getcwd()
        self.path = f'{self.adr}\\data_analysis\\'
        if self.option == 1:
            os.chdir('BS4')
            from BS4.ac_spider import spider_main
            ac = spider_main(self.num)
            ac.spider()
            self.team_path = join(self.adr, 'BS4\Data_Rank\Rank{}.xlsx'.format(self.num))
            self.member_path = join(self.adr, 'BS4\Data_Member\Member{}.xlsx'.format(self.num))
        else:
            os.chdir('Design\spiders')
            from Design.spiders.run import spider_main
            ac = spider_main(self.num)
            ac.spider()
            self.team_path = join(self.adr, 'Design\spiders\data_Rank\Rank{}.xlsx'.format(self.num))
            self.member_path = join(self.adr, 'Design\spiders\data_member\member{}.xlsx'.format(self.num))

        os.chdir(self.adr)
        return

    #导入数据
    def import_Data(self):
        print('Importing the data now!')
        print('waiting...')
        # 团队信息导入
        wb = load_workbook(self.team_path)
        try:
            rk = wb['Rank']
            for row in range(3, rk.max_row + 1):
                try:
                    self.team_rank.append(eval(rk['A' + str(row)].value.strip()))
                    self.team_school.append(rk['C' + str(row)].value.strip())
                    self.team_rating.append(eval(rk['E' + str(row)].value.strip()))
                    self.team_rating_contest.append(sum([eval(i) for i in rk['G' + str(row)].value.strip().split('/')]))
                    self.team_contest.append(eval(rk['H' + str(row)].value.strip()))
                    self.team_member.append(int(rk['J'+str(row)].value.strip()))
                    self.team_ac_summit.append(eval(rk['L' + str(row)].value) * 100 // eval(rk['M' + str(row)].value))
                    self.team_ac_rate.append(eval(rk['N' + str(row)].value[:-1]) * 100 // 100)
                except Exception:
                    print('Import1 Error!')
                    pass
        except Exception:
            print('Import1 Error!')
            pass
        wb.save(self.team_path)

        # 成员信息导入
        wb = load_workbook(self.member_path)
        try:
            info = wb['Information']
            for row in range(3, info.max_row + 1):
                try:
                    if info['B' + str(row)].value == ' 暂无':
                        self.member_rating.append(0)
                    else:
                        self.member_rating.append(eval(info['B' + str(row)].value.strip()))
                    self.member_rank.append(eval(info['C' + str(row)].value.strip()))
                    self.member_rating_contest.append(sum([eval(i) for i in info['D' + str(row)].value.strip().split('/')]))
                    self.member_contest.append(eval(info['E' + str(row)].value.strip()))
                    if eval(info['J' + str(row)].value) > 0:
                        self.member_ac_summit.append(eval(info['I' + str(row)].value) * 100 // eval(info['J' + str(row)].value))
                    if eval(info['K' + str(row)].value[:-1]) > 0:
                        self.member_ac_rate.append(eval(info['K' + str(row)].value[:-1]) * 100 // 100)
                except Exception:
                    print('Import2 Error!', row)
                    pass
        except Exception:
            print('Import2 Error!')
            pass
        wb.save(self.member_path)
        return

    #分布图处理
    def distribution(self, x):
        dict = {}
        if isinstance(x, str):
            x = [eval(i) for i in x]
        x.sort(key=lambda x: x)
        for key in x:
            if key in dict:
                dict[key] = dict[key] + 1
            else:
                dict[key] = 1
        x = list(dict.keys())
        y = list(dict.values())
        return x, y

    # 二次函数的标准形式
    def func(self,params, x):
        a, b, c = params
        return a * x * x + b * x + c

    # 误差函数，即拟合曲线所求的值与实际值的差
    def error(self,params, x, y):
        return self.func(params, x) - y

    # 对参数求解
    def slovePara(self,X,Y):
        p0 = [10, 10, 10]
        Para = leastsq(self.error, p0, args=(X, Y))
        return Para

    def fit(self,X,Y):
        Para = self.slovePara(np.array(X), np.array(Y))
        a, b, c = Para[0]
        z = [a * x * x + b * x + c for x in X]
        return z

    # 通过率分布，通过题数与提交题数比分布 团队/个人
    def image1(self, is_team):
        if is_team:
            name = 'team'
            x, y = self.distribution(self.team_ac_rate)
            u, v = self.distribution(self.team_ac_summit)
        else:
            name = 'member'
            x, y = self.distribution(self.member_ac_rate)
            u, v = self.distribution(self.member_ac_summit)
        trace1 = go.Scatter(
            x=x, y=y, name='真实通过率', xaxis='x1', yaxis='y1'
        )
        trace2 = go.Bar(
            x=x, y=y, name='真实通过率', xaxis='x2', yaxis='y2'
        )
        trace3 = go.Bar(
            x=u, y=v, name='对比通过率', xaxis='x2', yaxis='y2'
        )
        trace4 = go.Scatter(
            x=u, y=v, name='对比通过率', xaxis='x3', yaxis='y3'
        )
        trace5 = go.Scatter(
            x=x, y=y, name='真实通过率', xaxis='x4', yaxis='y4'
        )
        trace6 = go.Scatter(
            x=u, y=v, name='对比通过率', xaxis='x4', yaxis='y4'
        )
        data = [trace1, trace2, trace3, trace4, trace5, trace6]
        layout = go.Layout(
            title=f'{name} comparion',
            xaxis1=dict(
                domain=[0, 0.45], anchor='y1', ticksuffix='%'
            ),
            yaxis1=dict(
                domain=[0, 0.45], anchor='x1',
            ),
            xaxis2=dict(
                domain=[0, 0.45], anchor='y2', ticksuffix='%'
            ),
            yaxis2=dict(
                domain=[0.55, 1], anchor='x2'
            ),
            xaxis3=dict(
                domain=[0.5, 1], anchor='y3', ticksuffix='%'
            ),
            yaxis3=dict(
                domain=[0.55, 1], anchor='x3'
            ),
            xaxis4=dict(
                domain=[0.5, 1], anchor='y4', ticksuffix='%'
            ),
            yaxis4=dict(
                domain=[0, 0.45], anchor='x4'
            )
        )
        fig = go.Figure(data=data, layout=layout)
        if not os.path.exists(f'data_analysis\img{self.num}'):
            os.makedirs(f'data_analysis\img{self.num}')
        plot(fig, filename=f'data_analysis/img{self.num}/ac_img{name}.html')  # 斜杠不为反

    # 学历分布，学校分布 团队
    def image2(self):
        label = {'大学': 0, '中学': 0, '小学': 0, '其他': 0}
        school = {}
        member_num = {}
        for item in self.team_school:
            if item[-2:] in label:
                label[item[-2:]] = label[item[-2:]] + 1
            else:
                label['其他'] = label['其他'] + 1
            if item in school:
                school[item] = school[item] + 1
            else:
                school[item] = 1
        school = sorted(school.items(),key = lambda x:x[1],reverse=True)
        for item in self.team_member:
            if item in member_num:
                member_num[item] = member_num[item] + 1
            else:
                member_num[item] = 1
        u = [school[i][0] for i in range(min(20,len(school)))]
        v = [school[i][1] for i in range(min(20,len(school)))]
        x = list(label.keys())
        y = list(label.values())
        z = list(member_num.keys())
        w = list(member_num.values())
        fig = make_subplots(
            rows=2,
            cols=2,
            specs=[[{'type': 'domain'},
                    {'type': 'domain'}],
                    [{'type': 'domain'},
                    {'type': 'Scatter'}]]
        )
        trace1 = go.Pie(
            labels=x,values=y,name='学历分布'
        )
        trace2 = go.Scatter(
            x=u,y=v,name='school'#,xaxis='x',yaxis='y'
        )
        trace3 = go.Pie(
            labels=u,values=v,name='school'
        )
        trace4 = go.Pie(
            labels=z,values=w,name='member num'
        )
        fig.add_trace(trace1,1,1)
        fig.add_trace(trace3,1,2)
        fig.add_trace(trace4,2,1)
        # fig.update_traces(hole=.4, hoverinfo="label+percent+name")
        fig.update_layout(
            title_text="学校与学历分布",  # 图形的名字
            # 给甜甜圈添加注解
            # annotations=[dict(text='职业分布', x=0, y=0, font_size=10, showarrow=False),
            #              dict(text='学校分布', x=5, y=5, font_size=10, showarrow=False)]
        )
        fig.add_trace(trace2,2,2)
        if not os.path.exists(f'data_analysis\img{self.num}'):
            os.makedirs(f'data_analysis\img{self.num}')
        plot(fig,filename=f'data_analysis/img{self.num}/school_img.html')
        return

    # rating_contest与contest关系 团队and个人
    def image3(self):
        lx = ['rating赛','非rating赛']
        x = [sum(self.team_rating_contest),sum(self.team_contest)-sum(self.team_rating_contest)]
        y = [sum(self.member_rating_contest),sum(self.member_contest)-sum(self.member_rating_contest)]
        fig = make_subplots(
            rows=1,
            cols=2,
            specs = [
                [{'type':'domain'},{'type':'domain'}]
            ]
        )
        trace1 = go.Pie(
            labels=lx,values=x,name='team'
        )
        trace2 = go.Pie(
            labels=lx,values=y,name='member'
        )
        fig.add_trace(trace1,1,1)
        fig.add_trace(trace2,1,2)
        fig.update_traces(hole=.4, hoverinfo="label+percent+name")
        fig.update_layout(
            title_text="比赛分类",  # 图形的名字
            # 给甜甜圈添加注解
            annotations=[dict(text='团队', x=0.20, y=0.5, font_size=25, showarrow=False),
                         dict(text='个人', x=0.80, y=0.5, font_size=25, showarrow=False)])
        if not os.path.exists(f'data_analysis\img{self.num}'):
            os.makedirs(f'data_analysis\img{self.num}')
        plot(fig,filename=f'data_analysis/img{self.num}/contest_img.html')

        return

    ##排名与通过率，通过次数关系 团队and个人
    def image4(self):
        rank1 = self.team_rank
        rating1 = self.team_rating
        rank2 = [item for item in self.member_rank if item < 100000]
        rating2 = [item for item in self.member_rating if item > 0]
        length = min(len(rank2),len(rating2))
        while len(rank2) > length:
            rank2.pop()
        while len(rating2) > length:
            rating2.pop()
        rt1 = self.fit(rank1,rating1)
        rt2 = self.fit(rank2,rating2)

        fig = make_subplots(
            rows=1,
            cols=3,
            specs=[
                [{'type':'Scatter'},{'type':'Scatter'},{'type':'Scatter'}]
            ]
        )
        trace1 = go.Scatter(
            x=rank1,y=rating1,name='team'
        )
        trace2 = go.Scatter(
            x=rank2,y=rating2,name='member'
        )
        trace3 = go.Scatter(
            x=rank1,y=rt1,name='team fit'
        )
        trace4 = go.Scatter(
            x=rank2,y=rt2,name='member fit'
        )
        fig.add_trace(trace1,1,1)
        fig.add_trace(trace2,1,1)
        fig.add_trace(trace1,1,2)
        fig.add_trace(trace3,1,2)
        fig.add_trace(trace2,1,3)
        fig.add_trace(trace4,1,3)
        fig.update_layout(
            title='rating与排名关系图',
        )
        if not os.path.exists(f'data_analysis\img{self.num}'):
            os.makedirs(f'data_analysis\img{self.num}')
        plot(fig,filename=f'data_analysis/img{self.num}/rating_rank.html')
        return


    def image_produce(self):
        print('*'*100)
        print('1.团队通过率分布图 2.个人通过率分布图 3.学校与学历分布图 4.rating赛与非rating赛分布 5.rating与排名关系图 6.所有图')
        print('可输入多个标号，如：1 2 3')
        print('*'*100)
        option = input('输入标号：').split()
        if '6' in option:
            self.image1(True)
            self.image1(False)
            self.image2()
            self.image3()
            self.image4()
        else:
            if '1' in option:
                self.image1(True)
            if '2' in option:
                self.image1(False)
            if '3' in option:
                self.image2()
            if '4' in option:
                self.image3()
            if '5' in option:
                self.image4()
        print('*'*100)
        if self.option == 1:
            print(f'已将所有团队数据导入\BS4\Data_Rank\Rank{self.num}.xlsx')
            print(f'已将所有个人数据导入\BS4\Data_Member\Member{self.num}.xlsx')
        else:
            print(f'已将所有团队数据导入\Design\spiders\data_Rank\Rank{self.num}.xlsx')
            print(f'已将所有个人数据导入\Design\spiders\data_member\member{self.num}.xlsx')
        print(f'已将所选图导入\data_analysis\img{self.num}文件夹，格式为html')
        print('*'*100)
        return


def main():
    if __name__ == '__main__':
        AP = FINAL()
        AP.import_Data()
        AP.image_produce()
        return


# 最终接口
main()
