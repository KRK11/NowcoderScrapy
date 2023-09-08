import os
from openpyxl import Workbook
from openpyxl.styles import Alignment


class DesignPipeline:
    def process_item(self, item, spider):
        return item


# 三个爬虫的管道，分别初始化，处理，储存数据

class RankSpiderSpiderPipeline:
    def __init__(self):
        self.url = 'https://ac.nowcoder.com/acm/contest/profile/'
        self.rank = []
        with open('search_page.txt', 'r', encoding='utf-8') as tp:
            self.k = tp.read().strip()
        with open(r'url_team\url_team{}.txt'.format(self.k), 'w', encoding='utf-8'): pass
        with open(r'url_member\url_member{}.txt'.format(self.k),'w',encoding='utf-8'): pass

    def process_item(self, item, spider):
        if item['team']:
            self.rank.append([item['rank'],item['name'],item['school'],item['description'],item['rating'],item['id']])
            with open(r'url_team\url_team{}.txt'.format(self.k),'a',encoding='utf-8') as fp:
                fp.write(self.url+item['id'].strip()+'\n')
        else:
            with open(r'url_member\url_member{}.txt'.format(self.k),'a',encoding='utf-8') as fp:
                fp.write(self.url+item['id'].strip()+'\n')
        return item

    def close_spider(self,spider):
        self.rank.sort(key = lambda x:eval(x[0]))
        with open(r'data_Rank\data_tmp.txt','w',encoding='utf-8') as fp:
            for item in self.rank:
                for i in item:
                    fp.write(i.replace('\n',' ')+'!$!')
                fp.write('\n')


class TeamSpiderSpiderPipeline:
    def __init__(self):
        with open('search_page.txt', 'r', encoding='utf-8') as tp:
            self.k = tp.read().strip()
        self.prefix = 'https://ac.nowcoder.com/acm/contest/profile/'
        self.rank1 = {}
        self.rank2 = {}
        self.rank3 = {}
        self.uid = []

    def process_item(self, item, spider):
        if not item['fans']:
            self.rank3[item['id']] = [' ' + i for i in item['rating_contest']]
            self.uid.extend(item['contest'])
        elif item['page']:
            self.rank1[item['id']] = [item['rating_contest'],item['contest'],item['fans'],item['member_num']]
        else:
            self.rank2[item['id']] = [item['rating_contest'],item['contest'],item['fans'],item['member_num']]
        return item

    def close_spider(self,spider):
        wb = Workbook()
        wb.create_sheet('Rank')
        rank = wb['Rank']
        rank.merge_cells('A1:Q1')
        rank['A1'] = 'Rank'
        rank['A1'].alignment = Alignment(horizontal='center', vertical='center')
        rank['A2'] = 'rank'
        rank['B2'] = 'name'
        rank['C2'] = 'school'
        rank['D2'] = 'description'
        rank['E2'] = 'rating'
        rank['F2'] = 'id'
        rank['G2'] = 'rating contest'
        rank['H2'] = 'contest'
        rank['I2'] = 'fans'
        rank['J2'] = 'member num'
        rank['K2'] = 'challenge'
        rank['L2'] = 'accept'
        rank['M2'] = 'summit'
        rank['N2'] = 'ac_rate'
        rank.merge_cells('O2:Q2')
        rank['O2'] = 'member'
        rank['O2'].alignment = Alignment(horizontal='center', vertical='center')
        rk = []
        with open(r'data_Rank\data_tmp.txt','r',encoding='utf-8') as fp:
            for item in fp:
                new = item.strip().split('!$!')
                rk.append(new[:-1])
        for item in rk:
            if item[5].strip() in self.rank1:
                item.extend(self.rank1[item[5].strip()])
            if item[5].strip() in self.rank2:
                item.extend(self.rank2[item[5].strip()])
            if item[5].strip() in self.rank3:
                item.extend(self.rank3[item[5].strip()])
            if len(item)>6:
                rank.append(item)
        wb.save('data_Rank\Rank{}.xlsx'.format(self.k))
        os.remove(r'data_Rank\data_tmp.txt')
        with open(r'url_member\url_member{}.txt'.format(self.k),'a',encoding='utf-8') as fp:
            for uid in self.uid:
                fp.write(self.prefix+str(uid)+'\n')


class MemberSpiderSpiderPipeline:
    def __init__(self):
        with open('search_page.txt', 'r', encoding='utf-8') as tp:
            self.k = tp.read().strip()
        self.rank1 = {}
        self.rank2 = {}

    def process_item(self, item, spider):
        if (not item['name']) and (not item['challenge']):
            pass
        elif item['name']:
            self.rank1[item['id']] = [item['name'],item['rating'],item['rank'],item['rating_contest'],item['contest'],item['attention'],item['fans']]
        else:
            self.rank2[item['id']] = [item['challenge'],item['accept'],item['summit'],item['ac_rate']]
        return item

    def close_spider(self,spider):
        wb = Workbook()
        wb.create_sheet('Information')
        info = wb['Information']
        info['A1'] = 'Information'
        info.merge_cells('A1:K1')
        info['A1'].alignment = Alignment(horizontal='center', vertical='center')
        info['A2'] = 'name'
        info['B2'] = 'rating'
        info['C2'] = 'rank'
        info['D2'] = 'rating_contest'
        info['E2'] = 'contest'
        info['F2'] = 'attention'
        info['G2'] = 'fans'
        info['H2'] = 'challenge'
        info['I2'] = 'accept'
        info['J2'] = 'summit'
        info['K2'] = 'ac_rate'
        rk = []
        for item in self.rank1:
            if item in self.rank2:
                rk.append(self.rank1[item]+self.rank2[item])
        rk.sort(key = lambda x:eval(x[2]))
        for item in rk:
            info.append(item)
        wb.save('data_member\member{}.xlsx'.format(self.k))